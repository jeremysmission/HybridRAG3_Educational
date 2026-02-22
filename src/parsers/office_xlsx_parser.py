# ============================================================================
# HybridRAG -- XLSX Parser (src/parsers/office_xlsx_parser.py)
# ============================================================================
#
# WHAT THIS FILE DOES (plain English):
#   Reads Excel (.xlsx) spreadsheet files and converts them into
#   searchable text. Each sheet becomes a labeled section, and each
#   row becomes a pipe-delimited line (like: "Part No | Description | Qty").
#
# WHY THIS MATTERS:
#   Parts lists, test results, configuration tables, and inventory
#   spreadsheets contain structured data that people need to search.
#   "What part number is the antenna?" should find the row in the
#   equipment spreadsheet even though it's not a PDF or Word doc.
#
# HOW IT WORKS:
#   1. Open the .xlsx file using openpyxl in read-only mode (low RAM)
#   2. Loop through every sheet (workbook tab)
#   3. Tag each sheet with [SHEET] SheetName for traceability
#   4. For each row, convert all cell values to strings
#   5. Join cells with " | " pipe delimiters
#   6. Skip entirely empty rows (all cells blank)
#   7. Return combined text + details (sheet count, row count)
#
# LIMITATIONS:
#   This is not "perfect Excel understanding" -- it does not interpret
#   formulas, charts, or conditional formatting. But it makes the
#   TEXT CONTENT of every cell searchable, which covers 90%+ of use cases.
#
# INTERNET ACCESS: NONE
# ============================================================================

from __future__ import annotations

from pathlib import Path
from typing import Tuple, Dict, Any


class XlsxParser:
    def parse(self, file_path: str) -> str:
        text, _ = self.parse_with_details(file_path)
        return text

    def parse_with_details(self, file_path: str) -> Tuple[str, Dict[str, Any]]:
        path = Path(file_path)
        details: Dict[str, Any] = {"file": str(path), "parser": "XlsxParser"}

        try:
            import openpyxl
        except Exception as e:
            details["error"] = f"IMPORT_ERROR: {type(e).__name__}: {e}"
            return "", details

        try:
            # data_only=True reads the computed VALUES of formulas (not the
            # formula text). read_only=True prevents loading the entire file
            # into memory at once (important for huge spreadsheets).
            wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)

            parts = []
            sheets = 0
            rows_emitted = 0

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheets += 1
                parts.append(f"[SHEET] {sheet_name}")

                for row in ws.iter_rows(values_only=True):
                    # Convert row values to strings, skip empty rows
                    vals = []
                    for v in row:
                        if v is None:
                            vals.append("")
                        else:
                            vals.append(str(v).strip())

                    if all(x == "" for x in vals):
                        continue

                    rows_emitted += 1
                    parts.append(" | ".join(vals))

            full = "\n".join(parts).strip()
            details["total_len"] = len(full)
            details["sheets"] = sheets
            details["rows_emitted"] = rows_emitted

            return full, details
        except Exception as e:
            details["error"] = f"RUNTIME_ERROR: {type(e).__name__}: {e}"
            return "", details
